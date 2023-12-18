from yandexfreetranslate import YandexFreeTranslate 
import requests
from openpyxl import load_workbook
import csv

ex_file = 'quotes.xlsx'
wb = load_workbook(ex_file)
ws = wb["Лист1"]

file = open("quotes.txt", "w", encoding='utf-8')

csv_file = open("qoutes.csv", "w", encoding='utf-8')
writer = csv.writer(csv_file)


try:
    for row in range(1, 11):
        response = requests.get("https://favqs.com/api/qotd")
        quote = response.json()["quote"]["body"]
        yt = YandexFreeTranslate(api = "ios")
        rus_quote = yt.translate("en", "ru", quote)    
        
        file.write(rus_quote)
        file.write('\n')

        cell = ws.cell(row=row, column=1)
        cell.value = rus_quote
        wb.save(ex_file)
                
        writer.writerow([rus_quote])     



except Exception as e:
    print(e)

finally:
    file.close()
    wb.close()
