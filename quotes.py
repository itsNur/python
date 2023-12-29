import requests
from openpyxl import load_workbook
import csv
# import pandas as pd
from yandexfreetranslate import YandexFreeTranslate

def fetch_quotes():
    "Fetch quotes from the API and translate them to russian"
    quotes = []
    response = requests.get("https://api.chucknorris.io/jokes/random")
    quote = response.json()["value"]
    yt = YandexFreeTranslate(api = "ios")
    rus_quote = yt.translate("en", "ru", quote)
    quotes.append(rus_quote)
    return quotes

def write_to_txt(quotes):
    "Write quotes to a text file"
    with open("quotes.txt", "w", encoding='utf-8') as file:
        for quote in quotes:
            file.write(quote)
            file.write('\n')

def write_to_excel(quotes):
    "Write quotes to an Excel file"
    ex_file = 'quotes.xlsx'
    wb = load_workbook(ex_file)
    ws = wb["Лист1"]
    for row, quote in enumerate(quotes, start=1):
        cell = ws.cell(row=row, column=1)
        cell.value = quote
        wb.save(ex_file)

    wb.close()

def write_to_csv(quotes):
    "Write quotes to a CSV file"
    with open("quotes.csv", "w", encoding='utf-8', newline='') as csv_file:
        writer = csv.writer(csv_file)
        for quote in quotes:
            writer.writerow([quote])

if __name__ == "__main__":
    for _ in range(10):
        quotes = fetch_quotes()
        write_to_txt(quotes)
        write_to_excel(quotes)
        write_to_csv(quotes)